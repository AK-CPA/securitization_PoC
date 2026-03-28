"""Read Excel and XML data with heuristic range detection."""
import re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ── Excel functions ──────────────────────────────────────────────────────────

def get_sheet_names(filepath: str) -> list[str]:
    """Return list of sheet names from an Excel workbook."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_sheet_data(filepath: str, sheet_name: str, cell_range: str | None = None) -> list[list]:
    """Read data from a specific sheet, optionally within a cell range.

    Returns a 2D array of cell values (preserving types).
    """
    wb = load_workbook(filepath, read_only=False, data_only=True)
    ws = wb[sheet_name]

    if cell_range:
        min_col, min_row, max_col, max_row = parse_range(cell_range)
    else:
        min_row = ws.min_row or 1
        max_row = ws.max_row or 1
        min_col = ws.min_column or 1
        max_col = ws.max_column or 1

    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)

    wb.close()
    return data


def read_full_sheet(filepath: str, sheet_name: str) -> tuple[list[list], int, int, int, int]:
    """Read the entire used range of a sheet.

    Returns (data, min_row, min_col, max_row, max_col).
    """
    wb = load_workbook(filepath, read_only=False, data_only=True)
    ws = wb[sheet_name]

    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)

    wb.close()
    return data, min_row, min_col, max_row, max_col


# ── XML functions ────────────────────────────────────────────────────────────

def get_xml_tables(filepath: str) -> list[str]:
    """Return a list of table names found in an XML file.

    Heuristic: treats each direct child element of root as a potential
    "table" (by its tag name), or looks for repeating sibling elements
    that form tabular data. Returns unique tag names that have multiple
    child records.
    """
    tree = ET.parse(filepath)
    root = tree.getroot()

    # Strip namespace prefixes for readability
    ns = _strip_ns(root)

    # Find repeating elements (tables) — children of root with sub-children
    tag_counts: dict[str, int] = {}
    for child in root:
        tag = _local_tag(child.tag)
        if len(child) > 0:  # has sub-elements (looks like a record/row)
            tag_counts[tag] = tag_counts.get(tag, 0) + 1

    # If root children themselves are records (all same tag), treat root as one table
    if len(tag_counts) == 1:
        return [list(tag_counts.keys())[0]]

    # If root has distinct sections with repeating children, list those
    table_names = []
    seen = set()
    for child in root:
        tag = _local_tag(child.tag)
        if tag in seen:
            continue
        seen.add(tag)
        # Check if this element has repeating sub-elements (a table of records)
        sub_counts: dict[str, int] = {}
        for sub in child:
            sub_tag = _local_tag(sub.tag)
            sub_counts[sub_tag] = sub_counts.get(sub_tag, 0) + 1
        # Consider it a table if any sub-tag repeats 2+ times
        if any(v >= 2 for v in sub_counts.values()):
            table_names.append(tag)
        elif tag_counts.get(tag, 0) >= 2:
            # The tag itself repeats at root level — it's a flat table
            table_names.append(tag)

    if not table_names:
        # Fallback: just list direct child tags
        for child in root:
            tag = _local_tag(child.tag)
            if tag not in table_names:
                table_names.append(tag)

    return table_names


def read_xml_table(filepath: str, table_name: str) -> list[list]:
    """Read a table from an XML file as a 2D array.

    Finds elements matching `table_name` and converts child elements
    to rows, with field names as the first (header) row.
    """
    tree = ET.parse(filepath)
    root = tree.getroot()

    # Find the matching section
    records = _find_xml_records(root, table_name)

    if not records:
        return []

    # Collect all field names across records to build columns
    all_fields: list[str] = []
    field_set: set[str] = set()
    for rec in records:
        for child in rec:
            tag = _local_tag(child.tag)
            if tag not in field_set:
                field_set.add(tag)
                all_fields.append(tag)

    if not all_fields:
        return []

    # Build 2D array: header row + data rows
    data = [all_fields]  # header
    for rec in records:
        row = []
        field_map = {_local_tag(child.tag): (child.text or "").strip() for child in rec}
        for field in all_fields:
            val = field_map.get(field, "")
            # Try to convert numeric strings
            row.append(_try_numeric(val))
        data.append(row)

    return data


def _find_xml_records(root: ET.Element, table_name: str) -> list[ET.Element]:
    """Find record elements for a given table name."""
    local_table = table_name

    # Case 1: root's children match the table name directly (flat list)
    matching = [child for child in root if _local_tag(child.tag) == local_table]
    if matching and len(matching[0]) > 0:
        # The matching elements ARE the records
        return matching

    # Case 2: a child of root named table_name contains repeating sub-elements
    for child in root:
        if _local_tag(child.tag) == local_table:
            # Return its children as records
            if len(child) > 0:
                return list(child)

    # Case 3: search deeper
    for elem in root.iter():
        if _local_tag(elem.tag) == local_table and len(elem) > 0:
            return list(elem)

    return []


def _local_tag(tag: str) -> str:
    """Strip XML namespace prefix from a tag."""
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _strip_ns(root: ET.Element) -> str:
    """Extract namespace URI if present."""
    tag = root.tag
    if tag.startswith("{"):
        return tag[1:tag.index("}")]
    return ""


def _try_numeric(val: str):
    """Try to convert a string to int or float, return original if not numeric."""
    if not val:
        return val
    # Remove common formatting
    cleaned = val.replace(",", "").replace("$", "").strip()
    try:
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except (ValueError, TypeError):
        return val


# ── Unified helpers ──────────────────────────────────────────────────────────

def get_file_sheets(filepath: str) -> list[str]:
    """Get sheet/table names for any supported file type."""
    if filepath.lower().endswith(".xml"):
        return get_xml_tables(filepath)
    return get_sheet_names(filepath)


def read_file_data(filepath: str, sheet_name: str, cell_range: str | None = None) -> list[list]:
    """Read data from any supported file type."""
    if filepath.lower().endswith(".xml"):
        return read_xml_table(filepath, sheet_name)
    return read_sheet_data(filepath, sheet_name, cell_range)


# ── Range parsing utilities ──────────────────────────────────────────────────

def parse_range(cell_range: str) -> tuple[int, int, int, int]:
    """Parse an Excel range string like 'B5:K30' into (min_col, min_row, max_col, max_row)."""
    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', cell_range.strip().upper())
    if not match:
        raise ValueError(f"Invalid range format: {cell_range}. Use Excel notation like B5:K30.")

    min_col = column_index_from_string(match.group(1))
    min_row = int(match.group(2))
    max_col = column_index_from_string(match.group(3))
    max_row = int(match.group(4))

    return min_col, min_row, max_col, max_row


def range_to_string(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    """Convert numeric range bounds back to Excel notation like 'B5:K30'."""
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def get_sheet_dimensions(filepath: str, sheet_name: str) -> tuple[int, int]:
    """Get the max row and max column of a sheet."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    wb.close()
    return max_row, max_col
