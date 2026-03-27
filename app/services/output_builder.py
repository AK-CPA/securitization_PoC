"""Generate formatted output Excel workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Any


# Style constants
HEADER_FONT = Font(bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SOURCE_FONT = Font(italic=True, color="808080")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MISMATCH_FONT = Font(color="9C0006")


def build_output_workbook(
    comparisons: list[dict],
    output_path: str,
) -> str:
    """Build the output Excel workbook with one tab per table comparison.

    Args:
        comparisons: list of dicts, each containing:
            - table_label: str
            - word_filename: str
            - excel_filename: str
            - excel_tab_name: str
            - excel_range: str
            - word_table: list[list[str]]
            - excel_data: list[list]
            - diff_grid: list[list]
            - status_grid: list[list[str]]
            - row_precisions: list[int]
        output_path: path to write the .xlsx file

    Returns the output_path.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for comp in comparisons:
        # Create tab name (max 31 chars for Excel)
        tab_name = comp["table_label"][:31]
        ws = wb.create_sheet(title=tab_name)

        current_row = 1

        # Section 1: Offering Document Table
        current_row = _write_section_header(ws, current_row, "OFFERING DOCUMENT TABLE")
        current_row = _write_source_info(ws, current_row, f"Source: {comp['word_filename']}")
        current_row += 1  # Blank spacer

        word_table = comp["word_table"]
        for row_data in word_table:
            for c_idx, val in enumerate(row_data):
                ws.cell(row=current_row, column=c_idx + 1, value=val)
            current_row += 1

        current_row += 1  # Blank spacer

        # Section 2: Supporting Excel Table
        current_row = _write_section_header(ws, current_row, "SUPPORTING EXCEL TABLE")
        source_info = f"Source: {comp['excel_filename']} / Sheet: {comp['excel_tab_name']} / Range: {comp['excel_range']}"
        current_row = _write_source_info(ws, current_row, source_info)
        current_row += 1  # Blank spacer

        excel_data = comp["excel_data"]
        row_precisions = comp["row_precisions"]
        for r_idx, row_data in enumerate(excel_data):
            precision = row_precisions[r_idx] if r_idx < len(row_precisions) else 2
            for c_idx, val in enumerate(row_data):
                cell = ws.cell(row=current_row, column=c_idx + 1)
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.value = round(val, precision)
                    cell.number_format = _number_format(precision)
                else:
                    cell.value = val
            current_row += 1

        current_row += 1  # Blank spacer

        # Section 3: Comparison (X - Y)
        current_row = _write_section_header(ws, current_row, "COMPARISON (X \u2212 Y)")
        current_row = _write_source_info(ws, current_row, "Offering Doc value minus Excel value")
        current_row += 1  # Blank spacer

        diff_grid = comp["diff_grid"]
        status_grid = comp["status_grid"]
        for r_idx, (diff_row, status_row) in enumerate(zip(diff_grid, status_grid)):
            precision = row_precisions[r_idx] if r_idx < len(row_precisions) else 2
            for c_idx, (diff_val, status) in enumerate(zip(diff_row, status_row)):
                cell = ws.cell(row=current_row, column=c_idx + 1)

                if status == "skipped":
                    # Copy text label from word table if available
                    if r_idx < len(word_table) and c_idx < len(word_table[r_idx]):
                        cell.value = word_table[r_idx][c_idx]
                elif status == "type_mismatch":
                    cell.value = "TYPE ERR"
                    cell.fill = MISMATCH_FILL
                    cell.font = MISMATCH_FONT
                elif status == "match":
                    cell.value = 0
                    cell.number_format = _number_format(precision)
                elif status == "mismatch":
                    cell.value = diff_val
                    cell.number_format = _number_format(precision)
                    cell.fill = MISMATCH_FILL
                    cell.font = MISMATCH_FONT
                elif status == "unmatched":
                    cell.value = "UNMATCHED"
                    cell.fill = MISMATCH_FILL
                    cell.font = MISMATCH_FONT

            current_row += 1

        # Auto-fit column widths
        _auto_fit_columns(ws)

    wb.save(output_path)
    return output_path


def _write_section_header(ws, row: int, text: str) -> int:
    """Write a section header row and return next row."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left")
    # Extend header fill across columns
    for c in range(2, 20):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    return row + 1


def _write_source_info(ws, row: int, text: str) -> int:
    """Write a source info row and return next row."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SOURCE_FONT
    return row + 1


def _number_format(precision: int) -> str:
    """Generate Excel number format string for given precision."""
    if precision == 0:
        return "#,##0"
    return "#,##0." + "0" * precision


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        if col_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
